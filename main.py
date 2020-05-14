# -*- coding: UTF-8 -*-
import os
from openpyxl import Workbook
from openpyxl import load_workbook

excel_file = "SDK测试用例.xlsx"
goal_md_file = "test_list.md"

def realExcelFileContent(excel_file_path, goal_md_file):
    if not os.path.exists(excel_file_path):
        print("excel文件不存在")
    
    wb = load_workbook(excel_file_path)
    sheetnames = wb.sheetnames
    remind_str = ""
    for item in sheetnames:
        remind_str = remind_str + "1: " + item + "\n"
    
    code = input("输入需要转换的表格序号:\n %s"%remind_str)
    sheet_name = sheetnames[int(code) - 1]
    ws = wb[sheet_name]
    tuple_rows = ws.rows
    
    md_content = ""
    line_num = 0
    for item in tuple_rows: 

        if line_num == 1: #书写md格式
            style_content = "|"
            for cell in item:
                style_content = style_content + ":-:|"
            md_content = md_content + style_content + "\n"
            
        #拼接每行的内容
        row_content = "|"
        for cell in item:
            value = "" if cell.value == None else cell.value
            value = value.replace('\n',"<br>")

            row_content = row_content + value + "|"
        md_content = md_content + row_content + "\n"
        line_num += 1

    #生成md文件
    cur_path = os.getcwd()
    goal_file_path = os.path.join(cur_path, goal_md_file)
    file_obj = open(goal_file_path, 'w')
    file_obj.write(md_content)
    file_obj.close()

if __name__ == "__main__":
    realExcelFileContent(excel_file, goal_md_file)